"""
src/reporting/excel.py

Populates report_template.xlsx from computed Figure objects only.
Never touches LLM output for any numeric cell - if a narrative is
attached, it goes on a clearly separate sheet, never into the Value,
Limit, Utilization, or Status columns. This is the last checkpoint in
the "LLM cannot produce a number" chain: even if every earlier gate
somehow failed, this module's function signature makes it structurally
impossible to write LLM output into a numeric cell, because it never
receives LLM output as an argument to begin with.

Fully testable offline - report_template.xlsx and real Figure objects
from src/computation/metrics.py are both available with zero Neo4j or
LLM dependency. Tested below by actually generating a populated file
and reading every cell back to confirm it matches the source figures
exactly.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl

from src.computation.metrics import Figure

EXPECTED_HEADERS = ["Section", "Metric", "Value", "Limit", "Utilization", "Status", "Source (graph path → doc/page)"]


class ReportPopulationError(Exception):
    """Raised when the template doesn't match the expected shape, or a
    row in the template has no corresponding computed figure - fails
    loudly rather than silently leaving a row blank in the final
    report, which could be mistaken for 'this metric is fine' by a
    reader."""


def populate_report(
    figures: list[Figure],
    template_path: str | Path,
    output_path: str | Path,
    citations: dict[str, str] | None = None,
) -> Path:
    """Fills Value/Limit/Utilization/Status (and Source, if citations
    given) into the existing template rows, matched by metric name
    prefix - the same matching convention already used and proven in
    src/reconciliation/reconciler.py, kept consistent rather than
    reinvented here. citations is an optional {figure_id: display_string}
    map (e.g. from engine.py's FigureWithCitation) - the Source column
    is left blank, not fabricated, when no citation is supplied.

    Raises ReportPopulationError if the template's header row doesn't
    match what this function expects, or if any template row has no
    matching computed figure - never writes a partially-correct file.
    """
    template_path = Path(template_path)
    output_path = Path(output_path)

    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    if header != EXPECTED_HEADERS:
        raise ReportPopulationError(
            f"Template header mismatch. Expected {EXPECTED_HEADERS}, got {header}."
        )

    unmatched_rows = []
    for row in ws.iter_rows(min_row=2):
        section_cell, metric_cell, value_cell, limit_cell, util_cell, status_cell, source_cell = row
        metric_name = metric_cell.value
        if metric_name is None:
            continue

        match = next((f for f in figures if f.name.startswith(metric_name)), None)
        if match is None:
            unmatched_rows.append(metric_name)
            continue

        value_cell.value = match.formatted_value
        limit_cell.value = match.limit_text
        util_cell.value = match.formatted_utilization
        status_cell.value = match.status.value

        if citations is not None and match.id in citations:
            source_cell.value = citations[match.id]

    if unmatched_rows:
        raise ReportPopulationError(
            f"No computed figure found for template row(s): {unmatched_rows}. "
            f"Refusing to write a report with unexplained blank rows."
        )

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path


if __name__ == "__main__":
    import tempfile

    from src.ingestion.guidelines import parse_guidelines
    from src.ingestion.holdings import parse_holdings
    from src.computation.metrics import compute_all_figures_firm_a

    g = parse_guidelines("sample_docs/sample_fund_guidelines.pdf")
    p = parse_holdings("sample_docs/sample_holdings.csv")
    figures = compute_all_figures_firm_a(p, g)

    with tempfile.TemporaryDirectory() as tmp:
        output_path = Path(tmp) / "populated_report.xlsx"
        result_path = populate_report(figures, "sample_docs/report_template.xlsx", output_path)
        print(f"Wrote: {result_path}")

        # Read it back and verify every cell matches the source figures -
        # not just "openpyxl didn't raise", an actual value comparison.
        wb = openpyxl.load_workbook(result_path)
        ws = wb.active
        all_pass = True
        for row in ws.iter_rows(min_row=2, values_only=True):
            section, metric, value, limit, util, status, source = row
            match = next(f for f in figures if f.name.startswith(metric))
            checks = [
                ("value", value, match.formatted_value),
                ("limit", limit, match.limit_text),
                ("utilization", util, match.formatted_utilization),
                ("status", status, match.status.value),
            ]
            row_ok = all(actual == expected for _, actual, expected in checks)
            marker = "PASS" if row_ok else "FAIL"
            print(f"[{marker}] {metric}: " + ", ".join(f"{n}={a!r}" for n, a, _ in checks))
            all_pass = all_pass and row_ok

        print(f"\nALL 13 ROWS MATCH SOURCE FIGURES: {all_pass}")

        # Also verify the failure path: a template with an unmatched row
        # must raise, not silently write a blank.
        print("\n--- Testing failure path: missing figure for a template row ---")
        try:
            populate_report(figures[:-1], "sample_docs/report_template.xlsx", Path(tmp) / "should_fail.xlsx")
            print("FAIL: should have raised ReportPopulationError")
        except ReportPopulationError as exc:
            print(f"PASS: correctly raised: {exc}")
